import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Legge il file dipendenti.txt
with open("assets/dipendenti.txt", "r") as file:
    lines = file.readlines()

# Crea una lista di tuple (nome, colore)
dipendenti = []
for line in lines:
    line = line.strip()
    if line:  # ignora righe vuote
        parts = line.split(',')
        nome = parts[0].strip()
        colore = parts[1].strip() if len(parts) > 1 else None
        dipendenti.append((nome, colore))

# Crea un nuovo workbook e imposta il primo foglio come "UNO"
wb = Workbook()
ws1 = wb.active
ws1.title = "UNO"

# Crea il secondo foglio "DUE"
ws2 = wb.create_sheet("DUE")

# Nel foglio UNO scrivi i nomi nella riga 1
for col_index, (nome, colore) in enumerate(dipendenti, start=1):
    cella = ws1.cell(row=1, column=col_index)
    cella.value = nome
    if colore:
        cella.fill = PatternFill(start_color=colore, end_color=colore, fill_type="solid")

# Nel foglio DUE scrivi i nomi nella riga 2
for col_index, (nome, colore) in enumerate(dipendenti, start=1):
    cella = ws2.cell(row=2, column=col_index)
    cella.value = nome
    if colore:
        cella.fill = PatternFill(start_color=colore, end_color=colore, fill_type="solid")

# Salva il file Excel
wb.save("assets/dipendenti.xlsx")
