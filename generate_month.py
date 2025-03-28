import json
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import datetime
import calendar
from datetime import date, timedelta

def get_week_days(num_settimana, anno):
    import datetime
    # Definisco l'ordine dei giorni della settimana
    giorni = ["LUN", "MAR", "MER", "GIO", "VEN", "SAB", "DOM"]
    
    # Ottengo il lunedì della settimana ISO (anche se la data potrebbe appartenere all'anno precedente)
    lunedi = datetime.date.fromisocalendar(anno, num_settimana, 1)
    
    # Creo il dizionario: per ogni giorno della settimana, prendo il giorno del mese (day)
    giorni_mese = {giorni[i]: (lunedi + datetime.timedelta(days=i)).day for i in range(7)}
    
    return giorni_mese

def get_weeks_in_month(mese, anno):
    """Restituisce le settimane ISO che hanno almeno un giorno nel mese specificato."""
    # Ottengo il primo e l'ultimo giorno del mese
    primo_giorno = datetime.date(anno, mese, 1)
    ultimo_giorno = datetime.date(anno, mese, calendar.monthrange(anno, mese)[1])
    
    # Ottengo le settimane ISO di inizio e fine mese
    prima_settimana = primo_giorno.isocalendar()[1]
    ultima_settimana = ultimo_giorno.isocalendar()[1]
    
    # Verifica se l'anno ha 53 settimane o solo 52
    ultimo_giorno_anno = datetime.date(anno, 12, 31)
    settimane_anno = ultimo_giorno_anno.isocalendar()[1]
    max_settimane = settimane_anno
    
    # Gestisci il caso di cambio anno (dicembre -> gennaio)
    if ultima_settimana < prima_settimana:
        # Per l'anno corrente, considera fino all'ultima settimana disponibile
        settimane_prima_parte = list(range(prima_settimana, max_settimane + 1))
        # Per l'anno successivo, parti dalla settimana 1
        settimane_seconda_parte = list(range(1, ultima_settimana + 1))
        return settimane_prima_parte + settimane_seconda_parte
    else:
        return list(range(prima_settimana, ultima_settimana + 1))

def setup_worksheet(ws, dipendenti):
    """Imposta lo stile base del foglio di lavoro."""
    # Imposta un bordo sottile da applicare alle celle
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # 1) Colonna A: etichetta "ORARIO" su righe 1-2
    ws.merge_cells("A1:A2")
    cell_orario = ws["A1"]
    cell_orario.value = "ORARIO"
    cell_orario.alignment = Alignment(horizontal="center", vertical="center")
    cell_orario.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    cell_orario.font = Font(color="FFFFFF", bold=True)
    
    # Imposta larghezza colonna A
    ws.column_dimensions["A"].width = 8
    
    return thin_border

def add_time_slots(ws, thin_border):
    """Aggiunge le fasce orarie al foglio di lavoro."""
    # Inserisci gli orari (8:00 -> 20:00 ogni 30 min) nella colonna A dalla riga 5
    start_time = datetime.datetime(2023, 1, 1, 8, 0)
    end_time = datetime.datetime(2023, 1, 1, 20, 0)
    delta = datetime.timedelta(minutes=30)
    
    # Calcola il numero di time slots
    time_slots = 0
    current = start_time
    while current <= end_time:
        time_slots += 1
        current += delta
    
    # Calcoliamo la riga finale in base al numero di orari
    first_time_row = 5
    last_time_row = first_time_row + time_slots - 1
    
    current = start_time
    row_idx = first_time_row
    while current <= end_time:
        ws.cell(row=row_idx, column=1).value = current.strftime("%H:%M")
        row_idx += 1
        current += delta
        
    return first_time_row, last_time_row

def ensure_argb_color(color_code):
    """
    Ensures that a color code is in the correct aRGB format.
    If it's a 6-character hex (RGB), adds 'FF' prefix for opacity.
    """
    if not color_code:
        return "FFFFFFFF"  # Default to white if no color provided
    
    # Remove any # prefix if present
    if color_code.startswith('#'):
        color_code = color_code[1:]
    
    # If it's a 6-character RGB code, add FF prefix for opacity
    if len(color_code) == 6:
        return "FF" + color_code
    
    # If it's already 8 characters, assume it's correct
    if len(color_code) == 8:
        return color_code
    
    # For any other format, return white
    return "FFFFFFFF"

def crea_foglio_settimanale(wb, settimana, anno, dipendenti=None):
    """Crea un foglio settimanale all'interno di un workbook esistente."""
    if dipendenti is None:
        # Carica i dati dal file impostazioni.json
        with open("impostazioni.json", "r") as f:
            settings = json.load(f)
        dipendenti = settings["dipendenti"]
    
    # Carica i tipi di ore dal file impostazioni.json
    with open("impostazioni.json", "r") as f:
        settings = json.load(f)
        tipi_di_ore = settings.get("tipiDiOre", {})
    
    # Crea un nuovo foglio
    ws = wb.create_sheet(title=f"Sett{settimana}")
    
    # Giorni da visualizzare nelle intestazioni (in ordine)
    giorni = ["LUN", "MAR", "MER", "GIO", "VEN", "SAB", "DOM"]
    
    # Prendo il giorno del mese per ogni giorno della settimana
    giorni_mese = get_week_days(settimana, anno)
    
    # Imposta lo stile base del foglio
    thin_border = setup_worksheet(ws, dipendenti)
    
    # Aggiungi gli orari
    first_time_row, last_time_row = add_time_slots(ws, thin_border)
    
    # Calcola il numero di dipendenti
    num_dipendenti = len(dipendenti)
    
    # Calculate max_col
    max_col = 1 + num_dipendenti * 7  # colonna A + 7 blocchi da num_dipendenti colonne ciascuno
    
    # Struttura delle colonne per i giorni
    for i in range(7):  # Includiamo la domenica (0..6)
        day_name = giorni[i]   # LUN, MAR, ...
        day_num = giorni_mese[day_name]
        
        start_col = 2 + num_dipendenti * i
        end_col = start_col + num_dipendenti - 1
        
        # Riga 1: celle unite con il numero del giorno
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col
        )
        cell_day_num = ws.cell(row=1, column=start_col)
        cell_day_num.value = str(day_num)
        cell_day_num.alignment = Alignment(horizontal="center", vertical="center")
        
        # Riga 2: celle unite con il nome del giorno
        ws.merge_cells(
            start_row=2, start_column=start_col,
            end_row=2, end_column=end_col
        )
        cell_day_name = ws.cell(row=2, column=start_col)
        cell_day_name.value = day_name
        cell_day_name.alignment = Alignment(horizontal="center", vertical="center")
        
        # Colore di sfondo azzurrino per i giorni pari
        if day_num % 2 == 0:
            fill_color = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
            # Imposta il colore del testo a nero
            day_num_color = "000000"
            day_name_color = "000000"
        else:
            fill_color = None
            # Imposta il colore del testo a rosso
            day_num_color = "FF0000"
            day_name_color = "FF0000"
        
        # Applica il colore del testo e il grassetto
        cell_day_num.font = Font(color=day_num_color, bold=True)
        cell_day_name.font = Font(color=day_name_color, bold=True)
        
        # Applica il bordo e l'eventuale sfondo
        for r in range(1, last_time_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                # Non applicare i bordi alla riga vuota (4) tranne sotto
                if r == 4:
                    cell.border = Border(bottom=Side(border_style="thin", color="000000"))
                # Sfondo per le righe degli orari
                if fill_color and r >= 5:
                    cell.fill = fill_color
        
        # Imposta larghezza delle colonne
        for c in range(start_col, end_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 6
    
    # Inserisci i nomi dei dipendenti - senza colorazione
    for i in range(7):
        day_name = giorni[i]
        start_col = 2 + num_dipendenti * i
        
        for j, dip in enumerate(dipendenti):
            col = start_col + j
            cell_emp = ws.cell(row=3, column=col)
            cell_emp.value = dip["nome"]
            cell_emp.alignment = Alignment(horizontal="center", vertical="center")
            
            # Rimosso il codice che applica il colore di sfondo
    
    # Aggiungi una riga vuota tra i nomi dei dipendenti e gli orari (solo bordo inferiore)
    for c in range(1, max_col + 1):
        ws.cell(row=4, column=c).border = Border(bottom=Side(border_style="thin", color="000000"))
        ws.cell(row=4, column=c).value = None  # Clear row 4
    
    # Applica i bordi alla colonna A (orari)
    for r in range(1, last_time_row + 1):
        cell = ws.cell(row=r, column=1)
        # Per la riga 4 (riga vuota) applica solo il bordo inferiore
        if r == 4:
            cell.border = Border(bottom=Side(border_style="thin", color="000000"))
        else:
            cell.border = thin_border
    
    # Riga di "chiusura" con bordi
    for c in range(1, max_col + 1):
        cell = ws.cell(row=last_time_row + 1, column=c)
        cell.border = Border(top=Side(border_style="thin", color="000000"))
    
    # Rimuovi la formattazione condizionale esistente e sostituiscila con regole reali di Excel
    for i in range(7):  # Per ogni giorno della settimana
        start_col = 2 + num_dipendenti * i
        
        for j, dip in enumerate(dipendenti):  # Per ogni dipendente
            col = start_col + j
            colore_dipendente = ensure_argb_color(dip.get("colore", "FFFFFF"))  # Normalizza il colore
            
            # Crea una regola di formattazione condizionale per "L" (Lavoro)
            l_rule = CellIsRule(
                operator="equal",
                formula=['"L"'],
                stopIfTrue=True,
                fill=PatternFill(start_color=colore_dipendente, end_color=colore_dipendente, fill_type="solid")
            )
            
            # Applica la regola alle celle nell'area degli orari per questa colonna
            cell_range = f"{get_column_letter(col)}{first_time_row}:{get_column_letter(col)}{last_time_row}"
            ws.conditional_formatting.add(cell_range, l_rule)
            
            # Aggiungi regole per ciascun tipo di ora dal file impostazioni
            for codice, info_tipo in tipi_di_ore.items():
                if "colore" in info_tipo:
                    colore_tipo_ore = ensure_argb_color(info_tipo["colore"])  # Normalizza il colore
                    tipo_rule = CellIsRule(
                        operator="equal",
                        formula=[f'"{codice}"'],
                        stopIfTrue=True,
                        fill=PatternFill(start_color=colore_tipo_ore, end_color=colore_tipo_ore, fill_type="solid")
                    )
                    ws.conditional_formatting.add(cell_range, tipo_rule)
    
    return ws

def crea_settimanale(settimana=23, anno=2025):
    """Crea un file Excel per una singola settimana."""
    # Crea una nuova cartella Excel
    wb = openpyxl.Workbook()
    
    # Rimuovi il foglio predefinito
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Crea il foglio settimanale
    crea_foglio_settimanale(wb, settimana, anno)
    
    # Salva il file risultante
    wb.save(f"Sett{settimana}.xlsx")

def crea_mensile(mese=6, anno=2025):
    """Crea un file Excel mensile con un foglio per ogni settimana del mese."""
    # Carica i dati dal file impostazioni.json
    with open("impostazioni.json", "r") as f:
        settings = json.load(f)
    dipendenti = settings["dipendenti"]
    
    # Nome dei mesi in italiano
    mesi_ita = [
        "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
        "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE"
    ]
    
    # Ottieni tutte le settimane che hanno almeno un giorno in questo mese
    settimane = get_weeks_in_month(mese, anno)
    
    # Crea un nuovo workbook
    wb = openpyxl.Workbook()
    
    # Rimuovi il foglio predefinito
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Crea un foglio per ogni settimana
    for settimana in settimane:
        crea_foglio_settimanale(wb, settimana, anno, dipendenti)
    
    # Nome del file
    nome_mese = mesi_ita[mese - 1]  # -1 perché gli indici partono da 0
    nome_file = f"TURNI {nome_mese} {anno}.xlsx"
    
    # Salva il file
    wb.save(nome_file)
    return nome_file

if __name__ == "__main__":
    # Puoi chiamare crea_settimanale() o crea_mensile() qui
    # crea_settimanale()  # Per creare un singolo file settimanale
    # crea_mensile()    # Per creare un file mensile con tutti i fogli
    # Crea tutti i mesi per l'anno 2025
    for i in range(1, 13):
        crea_mensile(i, 2025)

